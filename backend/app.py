"""
FastAPI application for Legal Tabular Review system.
Provides REST API endpoints for all core operations.
"""

import logging
from typing import Optional, List, Dict, Any
import os
from datetime import datetime, timezone
from uuid import uuid4
import aiofiles
import asyncio
from dotenv import load_dotenv

from fastapi import FastAPI, File, UploadFile, HTTPException, BackgroundTasks, Request
from fastapi.responses import JSONResponse
from fastapi.concurrency import run_in_threadpool
from fastapi.middleware.cors import CORSMiddleware
from starlette.middleware.base import BaseHTTPMiddleware
from pydantic import BaseModel
import time
from collections import defaultdict

from src.models.schema import (
    ProjectCreateRequest, ProjectUpdateRequest, ProjectResponse,
    DocumentUploadRequest, DocumentResponse, ComparisonTableResponse,
    ExtractionUpdateRequest, TaskStatusResponse, EvaluationMetrics,
    EvaluationReportResponse, FieldTemplateCreate, FieldTemplateResponse,
    FieldType, AnnotationCreateRequest, AnnotationUpdateRequest,
)
from src.storage.repository import DatabaseRepository
from src.services.service_orchestrator import (
    ProjectService, DocumentService, ExtractionService,
    ReviewService, ComparisonService, EvaluationService, TaskService,
    DiffService, AnnotationService, ReExtractionService,
)

# Setup logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s',
)
logger = logging.getLogger(__name__)

# Load environment variables
load_dotenv()


# ==================== ENTERPRISE MIDDLEWARE ====================

class RequestLoggingMiddleware(BaseHTTPMiddleware):
    """Logs every request with method, path, status, and duration."""

    async def dispatch(self, request: Request, call_next):
        start = time.time()
        request_id = str(uuid4())[:8]
        logger.info(f"[{request_id}] --> {request.method} {request.url.path}")
        try:
            response = await call_next(request)
            duration = (time.time() - start) * 1000
            logger.info(
                f"[{request_id}] <-- {response.status_code} "
                f"({duration:.1f}ms) {request.method} {request.url.path}"
            )
            response.headers["X-Request-ID"] = request_id
            response.headers["X-Response-Time"] = f"{duration:.1f}ms"
            return response
        except Exception as e:
            duration = (time.time() - start) * 1000
            logger.error(f"[{request_id}] ERROR ({duration:.1f}ms): {str(e)}")
            raise


class RateLimitMiddleware(BaseHTTPMiddleware):
    """Simple in-memory rate limiter per client IP."""

    def __init__(self, app, max_requests: int = 100, window_seconds: int = 60):
        super().__init__(app)
        self.max_requests = max_requests
        self.window = window_seconds
        self.requests: Dict[str, list] = defaultdict(list)

    async def dispatch(self, request: Request, call_next):
        # Skip rate limiting for health checks
        if request.url.path in ("/health", "/api/v1/health"):
            return await call_next(request)

        client_ip = request.client.host if request.client else "unknown"
        now = time.time()

        # Clean old entries
        self.requests[client_ip] = [
            t for t in self.requests[client_ip] if now - t < self.window
        ]

        if len(self.requests[client_ip]) >= self.max_requests:
            return JSONResponse(
                status_code=429,
                content={
                    "error": "Rate limit exceeded",
                    "detail": f"Max {self.max_requests} requests per {self.window}s",
                },
            )

        self.requests[client_ip].append(now)
        response = await call_next(request)
        response.headers["X-RateLimit-Limit"] = str(self.max_requests)
        response.headers["X-RateLimit-Remaining"] = str(
            self.max_requests - len(self.requests[client_ip])
        )
        return response

# Initialize FastAPI app
app = FastAPI(
    title="Legal Tabular Review API",
    description="System for extracting key fields from legal documents and presenting them in structured tables",
    version="1.0.0"
)

# Add CORS middleware
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# Add enterprise middleware
app.add_middleware(RequestLoggingMiddleware)
app.add_middleware(RateLimitMiddleware, max_requests=200, window_seconds=60)

# Initialize database and services
DATABASE_URL = os.getenv("DATABASE_URL", "sqlite:///./legal_review.db")
repo = DatabaseRepository(DATABASE_URL)

project_service = ProjectService(repo)
document_service = DocumentService(repo)
extraction_service = ExtractionService(repo)
review_service = ReviewService(repo)
comparison_service = ComparisonService(repo)
evaluation_service = EvaluationService(repo)
task_service = TaskService(repo)
diff_service = DiffService(repo)
annotation_service = AnnotationService(repo)
re_extraction_service = ReExtractionService(repo)

# Global lock for document ingestion to prevent SQLite concurrency issues
ingest_lock = asyncio.Lock()


# ==================== HEALTH CHECK ====================

@app.get("/health")
async def health_check():
    """Health check endpoint."""
    return {
        "status": "healthy",
        "timestamp": datetime.now(timezone.utc).isoformat(),
    }


# ==================== PROJECT ENDPOINTS ====================

@app.post("/projects", response_model=ProjectResponse)
async def create_project(request: ProjectCreateRequest):
    """Create a new project."""
    try:
        project = project_service.create_project(
            name=request.name,
            description=request.description,
            field_template_id=request.field_template_id,
        )
        return project
    except Exception as e:
        logger.error(f"Error creating project: {str(e)}")
        raise HTTPException(status_code=400, detail=str(e))


@app.get("/projects/{project_id}", response_model=ProjectResponse)
async def get_project(project_id: str):
    """Get project information."""
    try:
        project = project_service.get_project_info(project_id)
        if not project:
            raise HTTPException(status_code=404, detail="Project not found")
        return project
    except Exception as e:
        logger.error(f"Error getting project: {str(e)}")
        raise HTTPException(status_code=400, detail=str(e))


@app.get("/projects")
async def list_projects(skip: int = 0, limit: int = 100):
    """List all projects."""
    try:
        projects = project_service.list_projects(skip, limit)
        return {
            "projects": projects,
            "total": len(projects),
        }
    except Exception as e:
        logger.error(f"Error listing projects: {str(e)}")
        raise HTTPException(status_code=400, detail=str(e))


@app.put("/projects/{project_id}", response_model=ProjectResponse)
async def update_project(project_id: str, request: ProjectUpdateRequest):
    """Update project."""
    try:
        project = project_service.update_project(
            project_id=project_id,
            name=request.name,
            description=request.description,
            field_template_id=request.field_template_id,
        )
        if not project:
            raise HTTPException(status_code=404, detail="Project not found")
        return project
    except Exception as e:
        logger.error(f"Error updating project: {str(e)}")
        raise HTTPException(status_code=400, detail=str(e))

@app.delete("/projects/{project_id}")
async def delete_project(project_id: str):
    """Delete project and all related data."""
    logger.info(f"Deleting project {project_id}")
    try:
        deleted = repo.delete_project(project_id)
        if not deleted:
            logger.warning(f"Project {project_id} not found for deletion")
            raise HTTPException(status_code=404, detail="Project not found")
        logger.info(f"Successfully deleted project {project_id}")
        return {"status": "deleted", "project_id": project_id}
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error deleting project: {str(e)}")
        raise HTTPException(status_code=400, detail=str(e))


# ==================== DOCUMENT ENDPOINTS ====================

@app.post("/projects/{project_id}/documents/upload", response_model=DocumentResponse)
async def upload_document(
    project_id: str,
    file: UploadFile = File(...),
):
    """Upload document to project."""
    try:
        # Save file to uploads directory
        upload_dir = os.path.join(os.path.dirname(__file__), "uploads")
        os.makedirs(upload_dir, exist_ok=True)
        
        # Sanitize filename
        safe_filename = os.path.basename(file.filename)
        file_path = os.path.join(upload_dir, safe_filename)
        
        # Handle duplicate filenames by appending timestamp
        if os.path.exists(file_path):
            name, ext = os.path.splitext(safe_filename)
            timestamp = datetime.now().strftime("%Y%m%d%H%M%S")
            file_path = os.path.join(upload_dir, f"{name}_{timestamp}{ext}")
            
        # Stream file to disk in chunks to avoid memory spikes
        CHUNK_SIZE = 1024 * 1024
        async with aiofiles.open(file_path, "wb") as f:
            while True:
                chunk = await file.read(CHUNK_SIZE)
                if not chunk:
                    break
                await f.write(chunk)

        # Ingest document
        # NOTE: Removed ingest_lock to allow parallel ingestion. 
        # DatabaseRepository now handles concurrency with retry_on_lock.
        logger.info(f"Invoking ingest_document for {file.filename} (Size: {os.path.getsize(file_path)} bytes)")
        
        start_time = datetime.now()
        try:
            document = await run_in_threadpool(
                document_service.ingest_document,
                project_id=project_id,
                filename=file.filename,
                file_path=file_path,
            )
            duration = (datetime.now() - start_time).total_seconds()
            logger.info(f"Ingestion successful for {file.filename} (Duration: {duration:.2f}s), Document ID: {document.get('id')}")
            return document
        except Exception as e:
            duration = (datetime.now() - start_time).total_seconds()
            logger.error(f"Ingestion failed for {file.filename} (Duration: {duration:.2f}s): {str(e)}")
            raise
    except Exception as e:
        logger.exception(f"Error uploading document {file.filename}: {str(e)}")
        raise HTTPException(status_code=400, detail=f"Upload failed for {file.filename}: {str(e)}")


@app.get("/projects/{project_id}/documents")
async def list_project_documents(project_id: str):
    """List documents in project."""
    try:
        documents = document_service.list_project_documents(project_id)
        return {
            "documents": documents,
            "total": len(documents),
        }
    except Exception as e:
        logger.error(f"Error listing documents: {str(e)}")
        raise HTTPException(status_code=400, detail=str(e))


# ==================== FIELD TEMPLATE ENDPOINTS ====================

@app.post("/field-templates", response_model=FieldTemplateResponse)
async def create_field_template(request: FieldTemplateCreate):
    """Create field template."""
    try:
        fields = [field.model_dump() for field in request.fields]
        template = repo.create_field_template(
            name=request.name,
            description=request.description,
            fields=fields,
        )
        return {
            'id': template.id,
            'name': template.name,
            'description': template.description,
            'version': template.version,
            'fields': template.fields,
            'created_at': template.created_at,
            'updated_at': template.updated_at,
            'is_active': template.is_active,
        }
    except Exception as e:
        logger.error(f"Error creating field template: {str(e)}")
        raise HTTPException(status_code=400, detail=str(e))


@app.get("/field-templates")
async def list_field_templates():
    """List field templates."""
    try:
        templates = repo.list_field_templates()
        return {
            "templates": [
                {
                    'id': t.id,
                    'name': t.name,
                    'version': t.version,
                    'fields_count': len(t.fields),
                    'created_at': t.created_at,
                }
                for t in templates
            ],
            "total": len(templates),
        }
    except Exception as e:
        logger.error(f"Error listing field templates: {str(e)}")
        raise HTTPException(status_code=400, detail=str(e))


# ==================== EXTRACTION ENDPOINTS ====================

class ExtractFieldsRequest(BaseModel):
    document_id: Optional[str] = None


@app.post("/projects/{project_id}/extract")
async def extract_fields(
    project_id: str,
    request: Optional[ExtractFieldsRequest] = None,
    background_tasks: BackgroundTasks = None,
):
    """Extract fields from documents."""
    try:
        # Get project and template
        project = repo.get_project(project_id)
        if not project:
            raise HTTPException(status_code=404, detail="Project not found")

        if not project.field_template_id:
            default_fields = [
                {
                    "name": "effective_date",
                    "display_name": "Effective Date",
                    "field_type": FieldType.DATE.value,
                    "description": "The effective date of the agreement",
                    "required": False,
                },
                {
                    "name": "parties",
                    "display_name": "Parties",
                    "field_type": FieldType.TEXT.value,
                    "description": "The parties involved in the agreement",
                    "required": False,
                },
                {
                    "name": "term",
                    "display_name": "Term",
                    "field_type": FieldType.TEXT.value,
                    "description": "The term or duration of the agreement",
                    "required": False,
                },
                {
                    "name": "governing_law",
                    "display_name": "Governing Law",
                    "field_type": FieldType.TEXT.value,
                    "description": "The governing law jurisdiction",
                    "required": False,
                },
                {
                    "name": "amount",
                    "display_name": "Amount",
                    "field_type": FieldType.CURRENCY.value,
                    "description": "The monetary amount referenced in the agreement",
                    "required": False,
                },
                {
                    "name": "payment_terms",
                    "display_name": "Payment Terms",
                    "field_type": FieldType.TEXT.value,
                    "description": "Payment terms or pricing schedule",
                    "required": False,
                },
                {
                    "name": "confidentiality",
                    "display_name": "Confidentiality",
                    "field_type": FieldType.TEXT.value,
                    "description": "Confidentiality obligations",
                    "required": False,
                },
                {
                    "name": "termination",
                    "display_name": "Termination",
                    "field_type": FieldType.TEXT.value,
                    "description": "Termination or cancellation terms",
                    "required": False,
                },
                {
                    "name": "indemnification",
                    "display_name": "Indemnification",
                    "field_type": FieldType.TEXT.value,
                    "description": "Indemnification obligations",
                    "required": False,
                },
                {
                    "name": "notice",
                    "display_name": "Notice",
                    "field_type": FieldType.TEXT.value,
                    "description": "Notice requirements",
                    "required": False,
                },
                {
                    "name": "jurisdiction",
                    "display_name": "Jurisdiction / Venue",
                    "field_type": FieldType.TEXT.value,
                    "description": "Jurisdiction and venue for disputes",
                    "required": False,
                },
                {
                    "name": "assignment",
                    "display_name": "Assignment",
                    "field_type": FieldType.TEXT.value,
                    "description": "Assignment and transferability terms",
                    "required": False,
                },
                {
                    "name": "force_majeure",
                    "display_name": "Force Majeure",
                    "field_type": FieldType.TEXT.value,
                    "description": "Force majeure clause",
                    "required": False,
                },
                {
                    "name": "dispute_resolution",
                    "display_name": "Dispute Resolution",
                    "field_type": FieldType.TEXT.value,
                    "description": "Arbitration/mediation or dispute resolution terms",
                    "required": False,
                },
                {
                    "name": "warranties",
                    "display_name": "Warranties",
                    "field_type": FieldType.TEXT.value,
                    "description": "Representations and warranties",
                    "required": False,
                },
                {
                    "name": "exclusivity",
                    "display_name": "Exclusivity",
                    "field_type": FieldType.TEXT.value,
                    "description": "Exclusivity obligations",
                    "required": False,
                },
                {
                    "name": "change_of_control",
                    "display_name": "Change of Control",
                    "field_type": FieldType.TEXT.value,
                    "description": "Change of control provisions",
                    "required": False,
                },
                {
                    "name": "amendment",
                    "display_name": "Amendment",
                    "field_type": FieldType.TEXT.value,
                    "description": "Amendment/modification terms",
                    "required": False,
                },
                {
                    "name": "severability",
                    "display_name": "Severability",
                    "field_type": FieldType.TEXT.value,
                    "description": "Severability clause",
                    "required": False,
                },
                {
                    "name": "waiver",
                    "display_name": "Waiver",
                    "field_type": FieldType.TEXT.value,
                    "description": "Waiver clause",
                    "required": False,
                },
                {
                    "name": "survival",
                    "display_name": "Survival",
                    "field_type": FieldType.TEXT.value,
                    "description": "Survival of obligations",
                    "required": False,
                },
                {
                    "name": "entire_agreement",
                    "display_name": "Entire Agreement",
                    "field_type": FieldType.TEXT.value,
                    "description": "Entire agreement clause",
                    "required": False,
                },
                {
                    "name": "counterparts",
                    "display_name": "Counterparts",
                    "field_type": FieldType.TEXT.value,
                    "description": "Counterparts clause",
                    "required": False,
                },
                {
                    "name": "audit_rights",
                    "display_name": "Audit Rights",
                    "field_type": FieldType.TEXT.value,
                    "description": "Right to audit records",
                    "required": False,
                },
                {
                    "name": "insurance",
                    "display_name": "Insurance",
                    "field_type": FieldType.TEXT.value,
                    "description": "Insurance requirements",
                    "required": False,
                },
                {
                    "name": "liability_cap",
                    "display_name": "Liability Cap",
                    "field_type": FieldType.CURRENCY.value,
                    "description": "Limitation of liability amount",
                    "required": False,
                },
                {
                    "name": "data_privacy",
                    "display_name": "Data Privacy",
                    "field_type": FieldType.TEXT.value,
                    "description": "Data protection and privacy terms",
                    "required": False,
                },
                {
                    "name": "non_solicitation",
                    "display_name": "Non-Solicitation",
                    "field_type": FieldType.TEXT.value,
                    "description": "Non-solicitation of employees/customers",
                    "required": False,
                },
                {
                    "name": "non_compete",
                    "display_name": "Non-Compete",
                    "field_type": FieldType.TEXT.value,
                    "description": "Non-competition restrictions",
                    "required": False,
                },
                {
                    "name": "subcontracting",
                    "display_name": "Subcontracting",
                    "field_type": FieldType.TEXT.value,
                    "description": "Rights to subcontract",
                    "required": False,
                },
                {
                    "name": "intellectual_property",
                    "display_name": "Intellectual Property",
                    "field_type": FieldType.TEXT.value,
                    "description": "IP ownership and licensing",
                    "required": False,
                },
                {
                    "name": "publicity",
                    "display_name": "Publicity",
                    "field_type": FieldType.TEXT.value,
                    "description": "Publicity and press release terms",
                    "required": False,
                },
            ]
            template = repo.create_field_template(
                name="Default Template",
                description="Auto-created template",
                fields=default_fields,
            )
            project = repo.update_project(project_id, field_template_id=template.id)
        else:
            template = repo.get_field_template(project.field_template_id)
            if not template:
                raise HTTPException(status_code=404, detail="Field template not found")

        field_definitions = template.fields

        # Create task
        task = task_service.create_task("extract", project_id)

        # Run extraction in background
        if background_tasks:
            background_tasks.add_task(
                _run_extraction,
                project_id,
                request.document_id if request else None,
                field_definitions,
                task['task_id'],
            )

        return {
            "task_id": task['task_id'],
            "status": "started",
            "message": "Extraction started in background",
        }

    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error extracting fields: {str(e)}")
        raise HTTPException(status_code=400, detail=str(e))


def _run_extraction(
    project_id: str,
    document_id: Optional[str],
    field_definitions: List[Dict[str, Any]],
    task_id: str,
):
    """Background task for field extraction."""
    try:
        task_service.repo.update_task(task_id, status='PROCESSING')

        if document_id:
            result = extraction_service.extract_fields_for_document(
                project_id, document_id, field_definitions
            )
        else:
            result = extraction_service.extract_all_documents(project_id, field_definitions)

        task_service.repo.update_task(
            task_id,
            status='COMPLETED',
            result=result,
        )
    except Exception as e:
        logger.error(f"Error in extraction background task: {str(e)}")
        task_service.repo.update_task(
            task_id,
            status='FAILED',
            error_message=str(e),
        )


# ==================== REVIEW ENDPOINTS ====================

@app.put("/extractions/{extraction_id}/review")
async def review_extraction(
    extraction_id: str,
    request: ExtractionUpdateRequest,
):
    """Review and update extraction."""
    try:
        result = review_service.update_extraction_review(
            extraction_id=extraction_id,
            status=request.status.value,
            manual_value=request.manual_value,
            reviewer_notes=request.reviewer_notes,
            reviewed_by=request.reviewed_by,
        )
        return result
    except Exception as e:
        logger.error(f"Error reviewing extraction: {str(e)}")
        raise HTTPException(status_code=400, detail=str(e))


@app.get("/projects/{project_id}/reviews/pending")
async def get_pending_reviews(project_id: str):
    """Get pending reviews for project."""
    try:
        reviews = review_service.get_pending_reviews(project_id)
        return {
            "reviews": reviews,
            "total": len(reviews),
        }
    except Exception as e:
        logger.error(f"Error getting pending reviews: {str(e)}")
        raise HTTPException(status_code=400, detail=str(e))


# ==================== COMPARISON TABLE ENDPOINTS ====================

@app.get("/projects/{project_id}/table")
async def get_comparison_table(project_id: str):
    """Get comparison table for project."""
    try:
        table = comparison_service.generate_comparison_table(project_id)
        return table
    except Exception as e:
        logger.error(f"Error generating comparison table: {str(e)}")
        raise HTTPException(status_code=400, detail=str(e))


@app.post("/projects/{project_id}/table/export-csv")
async def export_table_to_csv(project_id: str):
    """Export comparison table to CSV."""
    try:
        import csv
        import io

        table = comparison_service.generate_comparison_table(project_id)

        # Generate CSV
        output = io.StringIO()
        writer = csv.writer(output)

        # Headers
        headers = ["Field Name", "Field Type"]
        headers.extend([doc['filename'] for doc in table.get('documents', [])])
        writer.writerow(headers)

        # Rows
        for row in table.get('rows', []):
            row_data = [row['field_name'], row['field_type']]
            for doc in table.get('documents', []):
                doc_id = doc['id']
                result = row['document_results'].get(doc_id, {})
                value = result.get('extracted_value', 'N/A')
                row_data.append(value)
            writer.writerow(row_data)

        csv_content = output.getvalue()
        return {
            "format": "csv",
            "content": csv_content,
            "filename": f"legal_review_{project_id}.csv",
        }

    except Exception as e:
        logger.error(f"Error exporting table: {str(e)}")
        raise HTTPException(status_code=400, detail=str(e))


# ==================== EVALUATION ENDPOINTS ====================

@app.post("/projects/{project_id}/evaluate")
async def evaluate_project(
    project_id: str,
    evaluation_data: Dict[str, Any],
    background_tasks: BackgroundTasks = None,
):
    """Evaluate extraction quality."""
    try:
        # Create task
        task = task_service.create_task("evaluate", project_id)

        # Run evaluation in background
        if background_tasks:
            background_tasks.add_task(
                _run_evaluation,
                project_id,
                evaluation_data,
                task['task_id'],
            )

        return {
            "task_id": task['task_id'],
            "status": "started",
            "message": "Evaluation started in background",
        }

    except Exception as e:
        logger.error(f"Error starting evaluation: {str(e)}")
        raise HTTPException(status_code=400, detail=str(e))


def _run_evaluation(
    project_id: str,
    evaluation_data: Dict[str, Any],
    task_id: str,
):
    """Background task for evaluation."""
    try:
        task_service.repo.update_task(task_id, status='PROCESSING')

        items = evaluation_data.get('items', [])
        if items:
            for item in items:
                evaluation_service.evaluate_extraction(
                    project_id=project_id,
                    document_id=item.get('document_id'),
                    field_name=item.get('field_name'),
                    human_value=item.get('human_value'),
                )
            report = evaluation_service.generate_evaluation_report(project_id)
        else:
            # If no items provided, evaluate against all reviewed extractions
            report = evaluation_service.evaluate_project_reviews(project_id)

        task_service.repo.update_task(
            task_id,
            status='COMPLETED',
            result=report,
        )

    except Exception as e:
        logger.error(f"Error in evaluation background task: {str(e)}")
        task_service.repo.update_task(
            task_id,
            status='FAILED',
            error_message=str(e),
        )


@app.get("/projects/{project_id}/evaluation-report")
async def get_evaluation_report(project_id: str):
    """Get evaluation report for project."""
    try:
        report = evaluation_service.generate_evaluation_report(project_id)
        return report
    except Exception as e:
        logger.error(f"Error getting evaluation report: {str(e)}")
        raise HTTPException(status_code=400, detail=str(e))


# ==================== TASK ENDPOINTS ====================

@app.get("/tasks/{task_id}", response_model=TaskStatusResponse)
async def get_task_status(task_id: str):
    """Get async task status."""
    try:
        status = task_service.get_task_status(task_id)
        if not status:
            raise HTTPException(status_code=404, detail="Task not found")
        return status
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error getting task status: {str(e)}")
        raise HTTPException(status_code=400, detail=str(e))


# ==================== DIFF ENDPOINTS ====================

@app.get("/projects/{project_id}/diff")
async def get_project_diff(project_id: str):
    """Compute cross-document diff highlighting for a project."""
    try:
        diff_result = diff_service.compute_diff(project_id)
        return diff_result
    except Exception as e:
        logger.error(f"Error computing diff: {str(e)}")
        raise HTTPException(status_code=400, detail=str(e))


# ==================== ANNOTATION ENDPOINTS ====================

@app.post("/annotations")
async def create_annotation(request: AnnotationCreateRequest):
    """Create an annotation on an extraction."""
    try:
        annotation = annotation_service.create_annotation(
            extraction_id=request.extraction_id,
            comment_text=request.comment_text,
            annotated_by=request.annotated_by,
        )
        return annotation
    except Exception as e:
        logger.error(f"Error creating annotation: {str(e)}")
        raise HTTPException(status_code=400, detail=str(e))


@app.get("/extractions/{extraction_id}/annotations")
async def list_extraction_annotations(extraction_id: str):
    """List annotations for a specific extraction."""
    try:
        annotations = annotation_service.list_annotations_for_extraction(extraction_id)
        return {"annotations": annotations, "total": len(annotations)}
    except Exception as e:
        logger.error(f"Error listing annotations: {str(e)}")
        raise HTTPException(status_code=400, detail=str(e))


@app.get("/projects/{project_id}/annotations")
async def list_project_annotations(project_id: str):
    """List all annotations for a project."""
    try:
        annotations = annotation_service.list_annotations_for_project(project_id)
        return {"annotations": annotations, "total": len(annotations)}
    except Exception as e:
        logger.error(f"Error listing project annotations: {str(e)}")
        raise HTTPException(status_code=400, detail=str(e))


@app.put("/annotations/{annotation_id}")
async def update_annotation(annotation_id: str, request: AnnotationUpdateRequest):
    """Update an annotation."""
    try:
        annotation = annotation_service.update_annotation(annotation_id, request.comment_text)
        return annotation
    except ValueError as e:
        raise HTTPException(status_code=404, detail=str(e))
    except Exception as e:
        logger.error(f"Error updating annotation: {str(e)}")
        raise HTTPException(status_code=400, detail=str(e))


@app.delete("/annotations/{annotation_id}")
async def delete_annotation(annotation_id: str):
    """Delete an annotation."""
    try:
        deleted = annotation_service.delete_annotation(annotation_id)
        if not deleted:
            raise HTTPException(status_code=404, detail="Annotation not found")
        return {"status": "deleted", "annotation_id": annotation_id}
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error deleting annotation: {str(e)}")
        raise HTTPException(status_code=400, detail=str(e))


# ==================== RE-EXTRACTION ENDPOINTS ====================

@app.post("/projects/{project_id}/re-extract")
async def re_extract_project(
    project_id: str,
    background_tasks: BackgroundTasks = None,
):
    """Re-extract all fields for a project (deletes old extractions first)."""
    try:
        project = repo.get_project(project_id)
        if not project:
            raise HTTPException(status_code=404, detail="Project not found")

        template = None
        if project.field_template_id:
            template = repo.get_field_template(project.field_template_id)
        if not template:
            raise HTTPException(
                status_code=400,
                detail="No field template assigned to project. Assign a template first."
            )

        field_definitions = template.fields
        task = task_service.create_task("re-extract", project_id)

        if background_tasks:
            background_tasks.add_task(
                _run_re_extraction, project_id, field_definitions, task['task_id'],
            )

        return {
            "task_id": task['task_id'],
            "status": "started",
            "message": "Re-extraction started in background",
        }

    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error starting re-extraction: {str(e)}")
        raise HTTPException(status_code=400, detail=str(e))


def _run_re_extraction(
    project_id: str,
    field_definitions: List[Dict[str, Any]],
    task_id: str,
):
    """Background task for re-extraction."""
    try:
        task_service.repo.update_task(task_id, status='PROCESSING')
        result = re_extraction_service.re_extract_project(project_id, field_definitions)
        task_service.repo.update_task(task_id, status='COMPLETED', result=result)
    except Exception as e:
        logger.error(f"Error in re-extraction background task: {str(e)}")
        task_service.repo.update_task(task_id, status='FAILED', error_message=str(e))


# ==================== FIELD TEMPLATE UPDATE ENDPOINT ====================

@app.put("/field-templates/{template_id}")
async def update_field_template(template_id: str, request: FieldTemplateCreate):
    """Update a field template (creates new version)."""
    try:
        fields = [field.model_dump() for field in request.fields]
        template = repo.update_field_template(
            template_id=template_id,
            name=request.name,
            description=request.description,
            fields=fields,
        )
        if not template:
            raise HTTPException(status_code=404, detail="Field template not found")
        return {
            'id': template.id,
            'name': template.name,
            'description': template.description,
            'version': template.version,
            'fields': template.fields,
            'created_at': template.created_at,
            'updated_at': template.updated_at,
            'is_active': template.is_active,
        }
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error updating field template: {str(e)}")
        raise HTTPException(status_code=400, detail=str(e))


@app.get("/field-templates/{template_id}")
async def get_field_template(template_id: str):
    """Get a specific field template."""
    try:
        template = repo.get_field_template(template_id)
        if not template:
            raise HTTPException(status_code=404, detail="Field template not found")
        return {
            'id': template.id,
            'name': template.name,
            'description': template.description,
            'version': template.version,
            'fields': template.fields,
            'created_at': template.created_at,
            'updated_at': template.updated_at,
            'is_active': template.is_active,
        }
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error getting field template: {str(e)}")
        raise HTTPException(status_code=400, detail=str(e))


# ==================== EXCEL EXPORT ENDPOINT ====================

@app.post("/projects/{project_id}/table/export-excel")
async def export_table_to_excel(project_id: str):
    """Export comparison table to Excel (XLSX)."""
    try:
        import io
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

        table = comparison_service.generate_comparison_table(project_id)

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Legal Review Comparison"

        # Styles
        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_fill = PatternFill(start_color="2B579A", end_color="2B579A", fill_type="solid")
        header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
        thin_border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin'),
        )
        high_conf_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        med_conf_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
        low_conf_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

        # Headers
        headers = ["Field Name", "Field Type"]
        doc_names = []
        for doc in table.get('documents', []):
            doc_names.append(doc['filename'])
            headers.append(f"{doc['filename']} (Value)")
            headers.append(f"{doc['filename']} (Confidence)")

        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align
            cell.border = thin_border

        # Data rows
        for row_idx, row in enumerate(table.get('rows', []), 2):
            ws.cell(row=row_idx, column=1, value=row['field_name']).border = thin_border
            ws.cell(row=row_idx, column=2, value=str(row['field_type'])).border = thin_border

            col = 3
            for doc in table.get('documents', []):
                doc_id = doc['id']
                result = row['document_results'].get(doc_id, {})
                value = result.get('extracted_value', 'N/A')
                confidence = result.get('confidence_score', 0.0)

                value_cell = ws.cell(row=row_idx, column=col, value=value)
                value_cell.border = thin_border
                value_cell.alignment = Alignment(wrap_text=True)

                conf_cell = ws.cell(row=row_idx, column=col + 1, value=f"{confidence * 100:.0f}%")
                conf_cell.border = thin_border
                conf_cell.alignment = Alignment(horizontal="center")

                # Color-code confidence
                if confidence > 0.8:
                    conf_cell.fill = high_conf_fill
                elif confidence > 0.6:
                    conf_cell.fill = med_conf_fill
                elif confidence > 0:
                    conf_cell.fill = low_conf_fill

                col += 2

        # Auto-fit column widths
        for col_idx in range(1, len(headers) + 1):
            max_len = max(
                (len(str(ws.cell(row=r, column=col_idx).value or "")) for r in range(1, ws.max_row + 1)),
                default=10
            )
            ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = min(max_len + 4, 50)

        # Freeze header row
        ws.freeze_panes = "A2"

        # Save to bytes
        output = io.BytesIO()
        wb.save(output)
        excel_bytes = output.getvalue()

        import base64
        excel_b64 = base64.b64encode(excel_bytes).decode('utf-8')

        return {
            "format": "xlsx",
            "content_base64": excel_b64,
            "filename": f"legal_review_{project_id}.xlsx",
            "size_bytes": len(excel_bytes),
        }

    except ImportError:
        raise HTTPException(
            status_code=501,
            detail="openpyxl is required for Excel export. Install: pip install openpyxl"
        )
    except Exception as e:
        logger.error(f"Error exporting Excel: {str(e)}")
        raise HTTPException(status_code=400, detail=str(e))


# ==================== PROJECT EXTRACTIONS LISTING ====================

@app.get("/projects/{project_id}/extractions")
async def list_project_extractions(project_id: str):
    """List all extractions for a project (for annotation lookup)."""
    try:
        extractions = repo.list_extractions_by_project(project_id)
        return {
            "extractions": [
                {
                    'id': e.id,
                    'document_id': e.document_id,
                    'field_name': e.field_name,
                    'field_type': e.field_type.value if hasattr(e.field_type, 'value') else str(e.field_type),
                    'extracted_value': e.extracted_value,
                    'normalized_value': e.normalized_value,
                    'confidence_score': e.confidence_score,
                    'status': e.status.value,
                }
                for e in extractions
            ],
            "total": len(extractions),
        }
    except Exception as e:
        logger.error(f"Error listing extractions: {str(e)}")
        raise HTTPException(status_code=400, detail=str(e))


# ==================== ERROR HANDLERS ====================

@app.exception_handler(HTTPException)
async def http_exception_handler(request, exc):
    """Handle HTTP exceptions."""
    return JSONResponse(
        status_code=exc.status_code,
        content={
            "error": exc.detail,
            "status_code": exc.status_code,
        }
    )


if __name__ == "__main__":
    import uvicorn
    uvicorn.run(app, host="0.0.0.0", port=8000)
